import pandas as pd
import requests
import os
import time
import openpyxl
from datetime import datetime, timedelta

def get_monthly_revenue(year_month,market):
    print(f"Start to crawl {year_month}...")
    if market == "sii":
        url = f'https://mops.twse.com.tw/nas/t21/sii/t21sc03_{year_month}_0.html'
    elif market == "otc":
        url = f'https://mops.twse.com.tw/nas/t21/otc/t21sc03_{year_month}_0.html'

    folder_name = "StockRevenue"
    os.makedirs(folder_name, exist_ok=True)  # Create folder if it doesn't exist

    success = False
    attempts = 0
    while not success and attempts < 5:
        try:
            res = requests.get(url)
            res.encoding = 'big5'

            if res.status_code == 200:
                html_df = pd.read_html(res.text)
                df = pd.concat([df for df in html_df if df.shape[1] == 11])
                df.columns = df.columns.get_level_values(1)
                
                file_name = os.path.join(folder_name, f"stock_raw_{year_month}_{market}.xlsx")
                df.to_excel(file_name, index=False)
                print(f"Data for {year_month} saved to {file_name}")
                success = True
            else:
                raise Exception("Failed to fetch data")

        except Exception as e:
            print(f"Error occurred: {e}. Retrying in 5 seconds...")
            time.sleep(5)
            attempts += 1

    if not success:
        print(f"Failed to fetch data for {year_month} after {attempts} attempts.")
        
def find_missing_dates(monthly_dates, directory):
    # List all files in the directory
    files = os.listdir(directory)

    # Extract year_month from each file name and store in a set for efficient lookup
    file_dates = set()
    for file in files:
        if file.startswith('stock_raw_') and file.endswith('.xlsx'):
            # Extracting the date part from the filename
            date_part = file[len('stock_raw_'):-len('.xlsx')]
            file_dates.add(date_part)

    # Find out which dates from monthly_dates are not in file_dates
    missing_dates = [date for date in monthly_dates if date not in file_dates]

    return missing_dates

def calculate_dates():
    today = datetime.today()

    # Adjust the year and month for the URL format (year - 1911)
    latest_year = today.year - 1911
    latest_month = today.month

    # If today is before the 11th, go back one additional month
    if today.day < 11:
        latest_month -= 1

    # Adjust for year rollover
    if latest_month <= 0:
        latest_year -= 1
        latest_month += 12

    monthly_dates = []

    for i in range(1,14):
        year = latest_year
        month = latest_month - i
        if month <= 0:
            month += 12
            if month ==0:
                month = 12
                year -= 1
            year -= 1
            
        monthly_dates.append(f"{year}_{month}")

    return monthly_dates[::-1]  # Chronological order

def consolidate_data(monthly_dates, folder_name,market):
    all_data = {}

    for m in market:
        for date in monthly_dates:
            file_name = os.path.join(folder_name, f"stock_raw_{date}_{m}.xlsx")
            print(f"Processing file: {file_name}")

            if os.path.exists(file_name):
                workbook = openpyxl.load_workbook(file_name)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    company_code = row[0]
                    company_name = row[1]
                    revenue = row[2]

                    # Skip the row if the company name is '合計'
                    if company_name == '合計':
                        continue

                    if company_code not in all_data:
                        all_data[company_code] = {'CompanyName': company_name, f'{date}_Revenue': revenue}
                    else:
                        all_data[company_code][f'{date}_Revenue'] = revenue

            else:
                print(f"File not found: {file_name}")
                continue

    # Write combined data to a new Excel file
    write_combined_data_to_excel(monthly_dates,all_data)

def write_combined_data_to_excel(monthly_dates,data):
    
    file_name = os.path.join('StockRevenue', f"stock_revenue_statistics.xlsx")
    # Check if the file exists
    if os.path.exists(file_name):
        print(f"The file '{file_name}' exists.")
    else:
        print(f"The file '{file_name}' does not exist.")

    workbook = openpyxl.load_workbook(file_name)
    revenue_sheet = workbook["Revenue"]

    # Clear existing data (excluding the header row if it exists)
    revenue_sheet.delete_rows(1, revenue_sheet.max_row + 1)

    # Writing the header
    headers = ['CompanyCode', 'CompanyName'] + [f'{date}_Revenue' for date in monthly_dates]
    revenue_sheet.append(headers)

    # Writing the new data
    for company_code, info in data.items():

        # Start building the row with company code and company name
        row = [company_code, info['CompanyName']]
        #print(info)
        # Append revenue for each month, with debug prints
        for date in monthly_dates:
            revenue_key = f'{date}_Revenue'  # Construct the key for revenue
            revenue_value = info.get(revenue_key, '')  # Get the revenue value

            #Debug print: show the key and the retrieved value
            #print(f"Key: {revenue_key}, Value: {revenue_value}")

            # Append the revenue value to the row
            row.append(revenue_value)

        # Debug print: show the entire row before appending
        #print("Appending row:", row)

        # Append the row to the sheet
        revenue_sheet.append(row)

    workbook.save(file_name)
    print(f"Data saved to {file_name}")

if __name__ == '__main__':

    market = ["sii","otc"]
    #Output will be like ['111_12', '112_1', '112_2', '112_3', '112_4', '112_5', '112_6', '112_7', '112_8', '112_9', '112_10', '112_11', '112_12']
    monthly_dates = calculate_dates()

    # Get the current directory where the Python script is located
    current_directory = os.path.dirname(__file__)

    # Construct the relative path to the Excel file
    excel_file_path = os.path.join(current_directory, 'StockRevenue')
    
    missing_dates = find_missing_dates(monthly_dates, excel_file_path)
    if missing_dates:
        for m in market:
            # Crawl and save data for each month
            for date in missing_dates:
                get_monthly_revenue(date,m)
    else:
        print("No missing date.")

    consolidate_data(monthly_dates,'StockRevenue',market)
    print("Consolidated data saved to 'stock_revenue_statistics.xlsx'")
